"""
Excel export endpoint
"""
import io
import logging
from typing import List, Optional

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

logger = logging.getLogger(__name__)

router = APIRouter()

# Column headers for Excel export
EXPORT_COLUMNS = [
    "ID", "标题", "作者", "作者ID", "类型",
    "点赞数", "收藏数", "评论数", "分享数",
    "URL", "标签", "发布时间", "IP属地",
]


class ExportRequest(BaseModel):
    notes: List[dict]


@router.post("/export")
async def export_notes(
    notes: ExportRequest,
    filename: Optional[str] = Query(default="xhs_notes.xlsx"),
):
    if not filename:
        filename = "xhs_notes.xlsx"
    if not filename.endswith(".xlsx"):
        filename += ".xlsx"

    if not notes.notes:
        raise HTTPException(status_code=400, detail={"error": "no notes to export"})

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "笔记数据"

        # Style for header
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="FF5A5F")
        header_align = Alignment(horizontal="center", vertical="center")

        # Write headers
        for col, header in enumerate(EXPORT_COLUMNS, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Map NoteData fields to columns
        field_map = [
            "id", "title", "author", "author_id", "type",
            "liked", "collected", "commented", "shared",
            "url", "tags", "time", "ip_location",
        ]

        for row_idx, note in enumerate(notes.notes, start=2):
            for col_idx, field in enumerate(field_map, start=1):
                value = note.get(field, "")
                if field == "tags" and isinstance(value, list):
                    value = ", ".join(str(v) for v in value)
                elif value is None:
                    value = ""
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max_length + 2, 50)

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        filename_safe = "".join(c for c in filename if c.isalnum() or c in "._-")
        return StreamingResponse(
            iter([buffer.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename*=UTF-8\'\'{filename_safe}'
            },
        )

    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed. Run: pip install openpyxl")
    except Exception as e:
        logger.exception("Export failed")
        raise HTTPException(status_code=500, detail={"error": "export failed", "message": str(e)})
